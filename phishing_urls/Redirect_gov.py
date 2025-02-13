import time
import random
import openpyxl
from googlesearch import search

def is_whitelisted(url, whitelisted_domains):
    """Check if the URL belongs to a whitelisted domain."""
    for domain in whitelisted_domains:
        if domain in url:
            return True
    return False

def perform_search(query, whitelisted_domains=None, max_results=10):
    """Perform Google search and filter URLs based on whitelisted domains."""
    if whitelisted_domains is None:
        whitelisted_domains = []

    filtered_urls = []
    try:
        print("Starting Google search...")
        for url in search(query, num_results=max_results):
            print(f"Fetched URL: {url}")
            if not whitelisted_domains or not is_whitelisted(url, whitelisted_domains):
                filtered_urls.append(url)
                print(f"Added to filtered list: {url}")
            time.sleep(random.uniform(2, 5))  # Delay to avoid being flagged as a bot
    except Exception as e:
        print(f"Error during search: {e}")

    print(f"Total URLs fetched: {len(filtered_urls)}")
    return filtered_urls

def save_to_excel(data, filename):
    """Save the list of URLs to an Excel file."""
    if not data:
        print("No data to save to Excel.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Filtered URLs"

    # Write headers
    sheet.cell(row=1, column=1, value="Filtered URLs")

    # Write data
    for idx, url in enumerate(data, start=2):
        sheet.cell(row=idx, column=1, value=url)

    workbook.save(filename)
    print(f"Data saved to {filename}")

if __name__ == "__main__":
    # Google Dork Query - Modfy the domain and other keywords as needed
    query = ("porn | gaming | telegram | gambling | money | investment | stocks | crypto site:*.gov.in AND (inurl:.com | inurl:.net | inurl:.xyz | inurl:.ru | inurl:.tk | inurl:.ml | inurl:.cf | inurl:.ga | inurl:.gq | inurl:.info | inurl:.top | inurl:.biz | inurl:.pw | inurl:.cn | inurl:.co)")

    # Whitelisted domains (optional, add domains if needed)
    whitelisted_domains = []

    # Perform search
    results = perform_search(query, whitelisted_domains, max_results=100)

    # Save results to Excel
    output_file = "D:\\Desktop\\RedirectGOV\\filtered_urls.xlsx"
    save_to_excel(results, output_file)
